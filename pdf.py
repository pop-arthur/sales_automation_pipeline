#!/usr/bin/env python
# -*- coding: UTF-8 -*-
import csv
import json
import logging
from copy import copy
from datetime import datetime

from openpyxl.reader.excel import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Frame
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas


# email: 40 symbols max!
def form_pdf(file_name, name="Ivan Ivanov", phone="+79998887766", email="a@aaвввввввввddвввввввввввввввввdвввa.com",
             co_number="316", deliveryCond = "DDP г.Астана", data=[
            ['1', 'Лоток перфорированный 100х80 L2000, горячеоцинкованный', '35312HDZ', '8 087,00 KZT', 'Метр', '48',
             '388 176,00 KZT', '16', '#Н/Д', '10-15 раб дней'],
            ['2', 'Крышка на лоток с заземлением осн.100 L2000, горячеоцинкованная', '35512HDZ', '4 069,00 KZT', 'Метр',
             '48', '195 312,00 KZT', '16', '#Н/Д', '10-15 раб дней'],
        ]):
    file_name = file_name.replace(" ", "").replace(".", "")
    output_file = f'{file_name}.pdf'
    doc = SimpleDocTemplate(output_file, pagesize=A4)

    elements = []

    pdfmetrics.registerFont(TTFont('DejaVuSans', 'static/font/DejaVuSerif.ttf'))
    styles = getSampleStyleSheet()
    centered_style = ParagraphStyle(
        name='Centered',
        parent=styles['Normal'],
        fontName='DejaVuSans',
        fontSize=5,
        alignment=1,  # Center alignment (TA_CENTER)
        spaceAfter=12
    )
    left_style = ParagraphStyle(
        name='Left',
        parent=styles['Normal'],
        fontName='DejaVuSans',
        fontSize=5,
        alignment=0,  # Left alignment (TA_LEFT)
        spaceAfter=12
    )
    right_style = ParagraphStyle(
        name='Right',
        parent=styles['Normal'],
        fontName='DejaVuSans',
        fontSize=5,
        alignment=2,  # Right alignment (TA_RIGHT)
        spaceAfter=12
    )
    left_style_inv = ParagraphStyle(
        name='Left',
        parent=styles['Normal'],
        fontName='DejaVuSans',
        fontSize=5,
        alignment=0,  # Left alignment (TA_LEFT)
        spaceAfter=12,
        textColor=None
    )
    right_style_inv = ParagraphStyle(
        name='Right',
        parent=styles['Normal'],
        fontName='DejaVuSans',
        fontSize=5,
        alignment=2,  # Right alignment (TA_RIGHT)
        spaceAfter=12,
        textColor=None
    )
    big_bold_style = ParagraphStyle(
        name='BigBold',
        parent=styles['Normal'],
        fontName='DejaVuSans',
        fontSize=12,
        alignment=1,  # Center alignment (TA_CENTER)
        spaceAfter=12
    )

    p = Paragraph("<br/><br/><br/><br/><br/><br/>")
    elements.append(p)
    # Table data processing
    column_widths = [15, 120, 80, 60, 40, 25, 50, 60, 50, 60]
    style_normal = styles['Normal']
    style_normal.fontName = 'DejaVuSans'
    style_normal.wordWrap = 'NORMAL'
    # for i in range(len(data)):
    #     for j in range(len(data[i])):
    #         if len(str(data[i][j])) > 11:
    #             data[i][j] = str(data[i][j]).replace(" ", "\n")

    # Define the table style
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'DejaVuSans'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('rowHeights', (0, 0), (-1, -1), None)
    ])
    table_data = [['№', 'Наименование', 'Производитель\nили\nАртикул', 'Цена\nс учетом НДС', 'Ед.\nизм', 'К-во',
                   'Сумма\nс\nучетом НДС', 'Кратность\nупаковки', 'Наличие', 'Срок\nпоставки']]
    for row in data:
        arr = []
        for cell in row:
            arr.append(Paragraph(str(cell), styles['Normal']))
        table_data.append(arr)
    table = Table(table_data, style=table_style, colWidths=column_widths)
    elements.append(table)
    total = f"""ВСЕГО СТОИМОСТЬ ОБОРУДОВАНИЯ С УЧЕТОМ НДС: {count_items_price(items=data)}"""
    elements.append(Paragraph(total, big_bold_style))
    # Footer information
    footer_left = f"""
        <br/>
        Условия оплаты: 100% Предоплата<br/>
        Условие поставки: {deliveryCond}.<br/>
        Срок действия предложения: 5-7 календарных дней.<br/>
        Срок поставки оборудования: исчисляется с момента поступление предоплаты<br/>
        Контактное лицо поставщика: Калмаханов А.Г.<br/>
        E-mail: asb-top@mail.ru<br/>
        Моб: +7 708 425 62 82<br/>
        Тел: +7 7172 25-62-82<br/>
        Е. www.asb-top.kz<br/>
        С уважением,<br/>
        Калмаханов А.Г.
        """
    elements.append(Paragraph(footer_left, left_style))

    # Build the document
    doc.build(elements,
              onFirstPage=lambda canvas, doc: add_page_template_first(canvas, doc, co_number=co_number, name=name,
                                                                      phone=phone, email=email),
              onLaterPages=lambda canvas, doc: add_page_template(canvas, doc, co_number=co_number))


def add_page_template(canvas, doc, co_number="316"):
    canvas.saveState()
    pdfmetrics.registerFont(TTFont('DejaVuSans', 'static/font/DejaVuSerif.ttf'))
    canvas.setFont('DejaVuSans', 12)
    canvas.drawString(30, A4[1] - 50, "ТОО «АСБ-ТОП»")
    canvas.drawString(A4[0] - 230, A4[1] - 50, f"КП-№ {co_number}")
    canvas.restoreState()


def add_page_template_first(canvas, doc, co_number="as", name="ass", phone="dsdf", email="@"):
    canvas.saveState()
    pdfmetrics.registerFont(TTFont('DejaVuSans', 'static/font/DejaVuSerif.ttf'))
    canvas.setFont('DejaVuSans', 6)
    styles = getSampleStyleSheet()
    centered_style = ParagraphStyle(
        name='Centered',
        parent=styles['Normal'],
        fontName='DejaVuSans',
        fontSize=6,
        alignment=1,  # Center alignment (TA_CENTER)
        spaceAfter=12
    )
    left_style = ParagraphStyle(
        name='Centered',
        parent=styles['Normal'],
        fontName='DejaVuSans',
        fontSize=6,
        alignment=0,  # Center alignment (TA_CENTER)
        spaceAfter=12
    )
    big_bold_style = ParagraphStyle(
        name='Centered',
        parent=styles['Normal'],
        fontName='DejaVuSans',
        fontSize=12,
        alignment=1,  # Center alignment (TA_CENTER)
        spaceAfter=12
    )
    left_hren = """ТОО «АСБ-ТОП» Республика Казахстан,
        010000 г. Астана, Алматинский район,
        ул. Жанкент, д 155, офис 6.<br/>
        Тел: +7 777 5336031<br/>РНН 620200517080
        Р/сч KZ71998BTB0000323564 KZT;<br/>
        БИК TSESKZKA ; БИН 160540005119<br/>
        АО "Jýsan Bank"
        """
    paragraph = Paragraph(left_hren, centered_style)
    # Create a Frame to hold the paragraph (fixed-size box)
    frame_width = 150
    frame_height = 108
    frame_x = (17)
    frame_y = (700)  # Center vertically

    frame = Frame(frame_x, frame_y, frame_width, frame_height)

    # Add the paragraph to the frame
    frame.addFromList([paragraph], canvas)
    right_input = f"""Кому:{name}<br/> Тел:{phone}<br/>Почта:{email}<br/>КП-№{co_number}"""
    paragraph2 = Paragraph(right_input, centered_style)

    # Create a Frame to hold the paragraph (fixed-size box)
    frame_width = 200
    frame_height = 100
    frame_x = (400)
    frame_y = (690)  # Center vertically

    frame = Frame(frame_x, frame_y, frame_width, frame_height)
    frame.addFromList([paragraph2], canvas)
    company_name = Paragraph(
        f"""<b>ТОО "АСБ-ТОП"</b><br/><b>Коммерческое предложение</b><br/>от {datetime.date(datetime.now())} """,
        big_bold_style)
    # Create a Frame to hold the paragraph (fixed-size box)
    frame_width = 200
    frame_height = 55
    frame_x = (A4[0] - frame_width) / 2
    frame_y = (750)  # Center vertically

    frame = Frame(frame_x, frame_y, frame_width, frame_height, showBoundary=1)

    # Add the paragraph to the frame
    frame.addFromList([company_name], canvas)
    canvas.drawImage("static/images/logo.png", A4[0] - 120, A4[1] - 50, 50, 50)
    canvas.restoreState()


def form_excel(file_name, name, phone, email, co_number, data):
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

    # Create a logger instance
    logger = logging.getLogger(__name__)

    def copy_style(template_cell, new_cell):
        new_cell.font = copy(template_cell.font)
        new_cell.border = copy(template_cell.border)
        new_cell.fill = copy(template_cell.fill)
        new_cell.number_format = copy(template_cell.number_format)
        new_cell.protection = copy(template_cell.protection)
        new_cell.alignment = copy(template_cell.alignment)

    start_time = datetime.now()
    logger.info("Open file")
    # Load in the template
    wb = load_workbook('static/excel/commercial_offer_template.xlsx')
    # select sheet
    logger.info("Select sheet")
    sheet = wb[wb.sheetnames[0]]
    # set static data
    logger.info("Set static data")
    sheet["F17"] = name
    sheet["F18"] = phone
    sheet["F19"] = email
    sheet["F20"] = co_number

    # set items
    logger.info("Add rows")
    sheet.insert_rows(27, len(data) - 1)
    for i in range(len(data)):
        logger.info(f"Set row {i + 1} of {len(data)}")
        elem = data[i]
        row_number = 26 + i
        for j in range(10):
            new_cell = sheet.cell(row=row_number, column=j + 1)
            new_cell.value = elem[j]
            cell = sheet.cell(row=26, column=j + 1)
            copy_style(cell, new_cell)

    # set total price
    sheet.cell(row=28 + len(data), column=7).value = f"=SUM(G26:G{25 + len(data)})"
    sheet.merge_cells(start_row=28 + len(data), start_column=1, end_row=28 + len(data), end_column=3)
    sheet.cell(row=28 + len(data), column=1).value = " ВСЕГО СТОИМОСТЬ ОБОРУДОВАНИЯ С УЧЕТОМ НДС:"
    copy_style(sheet["D25"], sheet.cell(row=26 + len(data) + 2, column=1))
    copy_style(sheet["D25"], sheet.cell(row=26 + len(data) + 2, column=2))
    copy_style(sheet["D25"], sheet.cell(row=26 + len(data) + 2, column=3))

    logger.info("Save file")
    wb.save(f"{file_name}.xlsx")
    logger.info(f"Executed in {datetime.now() - start_time}")


# function to count cost of all items
def count_items_price(items):  # подсчёт цены всех товаров
    price = 0
    for item in items:
        if item[0] != "№":
            price += float(str(item[6]).replace("\n", "").replace("KZT", "").replace(",", ".").replace(" ", ""))
    return price


# function that combines formin pdf and csv, also prepare data for them
def form_files_from_list(products, fio, phone, email, coef, delcond,
                         is_form_pdf=True):  # Формирование КП на вход idишники -> на выход pdf'ка и csv
    items = []
    fio = fio.replace(" ", "")
    # jsontext = load_json()
    for i in range(len(products)):
        # print(products[i])
        toAppend = [i + 1, products[i]["name"].replace("Name: ", ""), products[i]["sku"].replace("SKU: ", ""),
                    float(products[i]["price"].split(" ")[1]) * float(coef),  products[i]["measure"], products[i]["amount"],
                    float(products[i]["price"].split(" ")[1]) * coef * int(products[i]["amount"]), "64",
                    products[i]["quantity"], products[i]["deliveryDate"]]
        items.append(toAppend)
    itForCSV = items.copy()
    if is_form_pdf:
        form_pdf(f"КП{fio + str(datetime.now().date())}", name=fio, phone=phone, email=email, deliveryCond = delcond, data=items)
    else:
        form_excel(f"КП{fio + str(datetime.now().date())}", name=fio, phone=phone, email=email, data=items, co_number="300 (temp)")
    # make_pdf(items=items, fileName=fio + '.pdf', name=fio, phone=phone, mail=email)



# function to find item in json by id
def find_in_json_by_id(findId,
                       jsonText=None):  # На вход: опционально json с продуктами и id для поиска -> на выход конкретный продукт из json'ая
    if (jsonText == None):
        jsonText = load_json()
    dic = json.loads(jsonText)
    products = dic["products"]
    for product in dic.get('products', []):
        if product['id'] == findId:
            return product
    return None


# function to find item in json by it's name
def find_in_json(findName, jsonText=None):
    if (jsonText == None):
        jsonText = load_json()
    dic = json.loads(jsonText)
    products = dic["products"]
    to_return = []
    for product in dic.get('products', []):
        if str(findName).lower() in str(product['name']).lower():
            to_return.append(product)
    return to_return


# function to find item in lsit by it's name
def find_item(findName, arr):
    to_return = []
    for product in arr:
        if str(findName).lower() in str(product['name']).lower() or str(findName).lower() in str(
                product['sku']).lower():
            to_return.append(product)
    return to_return


# function to load text from json file
def load_json():  # достаём json из файла
    file_path = "apiresp.json"
    with open(file_path, 'r', encoding='utf-8') as file:
        data = file.read().rstrip()
    jsonText = data
    jsonText = jsonText.replace("\n", "")
    return jsonText


# function to pars data from user. To put on satu.kz
def parsing_file(filename):
    # TODO make this function :)
    print("Asdsad")
    return


if __name__ == "__main__":
    # TODO rewrite
    # form_files_from_list([115900610, 109158468])  # test code. #delete before production
    # make_csv()
    data = [
        ['1', 'Лоток перфорированный 100х80 L2000, горячеоцинкованный', '35312HDZ', '8 087,00 KZT', 'Метр', '48',
         '388 176,00 KZT', '16', '#Н/Д', '10-15 раб дней'],
        ['2', 'Крышка на лоток с заземлением осн.100 L2000, горячеоцинкованная', '35512HDZ', '4 069,00 KZT', 'Метр',
         '48', '195 312,00 KZT', '16', '#Н/Д', '10-15 раб дней'],
        ['1', 'Лоток перфорированный 100х80 L2000, горячеоцинкованный', '35312HDZ', '8 087,00 KZT', 'Метр', '48',
         '388 176,00 KZT', '16', '#Н/Д', '10-15 раб дней'],
        ['2', 'Крышка на лоток с заземлением осн.100 L2000, горячеоцинкованная', '35512HDZ', '4 069,00 KZT', 'Метр',
         '48', '195 312,00 KZT', '16', '#Н/Д', '10-15 раб дней'],
        ['1', 'Лоток перфорированный 100х80 L2000, горячеоцинкованный', '35312HDZ', '8 087,00 KZT', 'Метр', '48',
         '388 176,00 KZT', '16', '#Н/Д', '10-15 раб дней'],
    ]
    form_pdf("CO1", data=data)
